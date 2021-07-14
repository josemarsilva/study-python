#
# filename    : lab-129-openpyxl.py
# Description : Open, read, write, manipulate Excel files
# Docs        : * https://openpyxl.readthedocs.io/en/stable/tutorial.html
# Requirements:
#               * pip install openpyxl
#

import openpyxl
from random import uniform

wb = openpyxl.load_workbook('file-09-excel-2007-365-workbook.xlsx')
ws_list = wb.sheetnames
for ws_name in ws_list:
    ws = wb[ws_name]
    print(f'+ {ws_name}:')
    row_num = 0
    for rows in ws.values:
        row_num += 1
        print(f'  {row_num}: {rows}')

# write some thing
ws_create = wb.create_sheet()

wb.save('file-10-excel-2007-365-workbook.xlsx')
